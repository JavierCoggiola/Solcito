from django.views.decorators.http import require_POST, require_GET
from django.shortcuts import render_to_response, render, redirect
from django.template import RequestContext
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from Solcito.forms import UploadForm
from Solcito.models import Student, Teacher, Marks, RegistrationD, trim, RegistrationS, Subject
import openpyxl
import xlrd

import datetime

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

@login_required(login_url='/login/')
def docentes(request):
    context = RequestContext(request)
    if request.user.groups.filter(name='Docente').exists():
        matriculas = RegistrationD.objects.filter(teacher=request.user.teacher)
        trimestres =  trim
        return render_to_response('docentes.html', {'matriculas': matriculas, 'trimestres' : trimestres}, context)
    else:
        return redirect('/admin')


def materias(request):
    context = RequestContext(request)
    if request.user.groups.filter(name='Docente').exists():
        materias = Subject.objects.filter(teachs__teacher=request.user.teacher)
        return render_to_response('materias.html', {'materias': materias}, context)
    else:
        return redirect('/admin')


def get_by_materia(request):
    context = RequestContext(request)
    if request.user.groups.filter(name='Docente').exists():
        materia = request.GET['materia']
        students = RegistrationS.objects.filter(curso__ofcurso= materia)
        for student in students:
            student.trimestres = []
            # por cada trimestre seteado filtramos las notas de los alumnos y se las vamos agregando
            for t in trim:
                trimestre = {}
                trimestre['notas'] = Marks.objects.filter(subject=materia, trim=t[0], reg=student)
                if len(trimestre['notas'])>0:
                    notas = [int(mark.nota) for mark in trimestre['notas']]
                    trimestre['final'] = sum(notas)/len(notas)
                student.trimestres.append(trimestre)
        return render_to_response('by_materia.html', {'students': students, 'trimestres':trim}, context)
    else:
        return redirect('/admin')



def excel(request):
    context = RequestContext(request)
    input_excel = request.FILES['docfile']
    materia = request.POST['materia']
    trimestre = request.POST['trimestre']
    book = xlrd.open_workbook(file_contents=input_excel.read())
    rows = []
    # leer hoja 1
    sh = book.sheet_by_index(0)
    print trimestre
    offset = 7
    if trimestre == '1':
        desde = 3
        hasta = 6
        columna = 7
    elif trimestre == '2':
        desde = 8
        hasta = 11
        columna = 12
    elif trimestre == '3':
        desde = 13
        hasta = 16
        columna = 17
    else:
        desde = 3
        hasta = 6
        columna = 7
    # fortuna row 15 col 2
    breaking = False
    for i, row in enumerate(range(sh.nrows)):
        if i <= offset:  # (Optionally) skip headers
            continue
        r = []
        for j, col in enumerate(range(sh.ncols)):
            #if (sh.cell_value(i, columna)) != "":
            if j<17:
                r.append(sh.cell_value(i, j))
            else:
                break
        if (sh.cell_value(i, columna)) == "":
            break
        rows.append(r)
    print rows
    # Guardar los datos del excel en la base de datos
    this_year = int(datetime.datetime.now().year)
    for i in rows:
        # Por cada alumno osea cada fila, pasamos por las columnas desde hasta como marcamos antes segun el trimestre
        print desde
        desdeAux = desde
        while desdeAux <=hasta:
            if str(i[desdeAux]) != "":
                newMark = Marks()
                try:
                    newMark.reg = RegistrationS.objects.get(student__dni=i[2], curso__cycle=this_year)
                    newMark.subject = Subject.objects.get(pk=materia)
                    print "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
                    print trimestre
                    newMark.trim = trimestre
                    newMark.nota = int(i[desdeAux])
                    newMark.save()
                    desdeAux += 1
                except IndexError:
                    print "error"
                    pass

    return render_to_response('docentes.html', {'rows': rows}, context)


def createModels(request):
    context = RequestContext(request)
    for i in rows:
        newalumn = Student(name=i[2], nota=i[7])
        newalumn.save()
    return render_to_response('index.html', {}, context)


@login_required(login_url='/login/')
def cleanExcel(request):
    context = RequestContext(request)
    if request.user.groups.filter(name='Docente').exists():
        subjects = Subject.objects.filter(teachs__in=request.user.teacher.ownerregistration.all())
        return render_to_response('clean_excel.html', {'subjects': subjects}, context)
    else:
        return redirect('/admin')


def getSubjectExcel(request):
    mat= request.GET.get('materia')
    materia = Subject.objects.get(pk=mat)
    materiaName = materia.name

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;')
    response['Content-Disposition'] = 'attachment; filename=nuevo_excel_{}.xlsx'.format(materiaName)

    title = Font(name='Arial', size=14, vertAlign="baseline", b=True, )
    subtitle14 = Font(name='Arial', size=14)
    subtitle12 = Font(name='Arial', size=12, b=True)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    alumnos = RegistrationS.objects.filter(curso=materia.curso)

    data = {
        'alumnos': alumnos,
    }

    workbook = Workbook()
    workbook.remove_sheet(workbook.active)

    ws = workbook.create_sheet()
    ws.title = "{}".format(materiaName)
    title_range = ws.merge_cells('A2:R2')
    subtitle_range = ws.merge_cells('A3:R3')

    primTrim = ws.merge_cells('D8:G8')
    SegTrim = ws.merge_cells('I8:L8')
    TerTrim = ws.merge_cells('N8:Q8')

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 13
    ws['A2'].font = title
    ws['A3'].font = subtitle14
    ws['A2'] = materiaName
    ws.row_dimensions[2].height = 40
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A3'] = materia.curso.cycle
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A5'].font = subtitle12
    ws['A5'] = "Curso: {}".format("7mo D")
    ws['A6'].font = subtitle12
    ws['A6'] = "Turno: {}".format("tarde")

    ws['A8'].border = thin_border
    ws['A8'] = "N de Orden"
    ws['A8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['B8'].border = thin_border
    ws['B8'] = "Nombre"
    ws['B8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['C8'].border = thin_border
    ws['C8'] = "D.N.I."
    ws['C8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['D8'].border = thin_border
    ws['D8'] = "1er Trim"
    ws['D8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['I8'].border = thin_border
    ws['I8'] = "2do Trim"
    ws['I8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['N8'].border = thin_border
    ws['N8'] = "3er Trim"
    ws['N8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['H8'].border = thin_border
    ws['H8'] = "Prom"
    ws['H8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['M8'].border = thin_border
    ws['M8'] = "Prom"
    ws['M8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['R8'].border = thin_border
    ws['R8'] = "Prom"
    ws['R8'].alignment = Alignment(horizontal='center', vertical='center')

    ws['R8'].border = thin_border

    count = 0
    number = 8
    # for matricula in alumno.getMatriculasAlumnos():
    for alumno in alumnos:
        count += 1
        orden = ws['A{}'.format(number + count)]
        name = ws['B{}'.format(number + count)]
        dni = ws['C{}'.format(number + count)]

        nota1prim = ws['D{}'.format(number + count)]
        nota1seg = ws['I{}'.format(number + count)]
        nota1ter = ws['N{}'.format(number + count)]

        nota2prim = ws['E{}'.format(number + count)]
        nota2seg = ws['J{}'.format(number + count)]
        nota2ter = ws['O{}'.format(number + count)]

        nota3prim = ws['F{}'.format(number + count)]
        nota3seg = ws['K{}'.format(number + count)]
        nota3ter = ws['P{}'.format(number + count)]

        nota4prim = ws['G{}'.format(number + count)]
        nota4seg = ws['L{}'.format(number + count)]
        nota4ter = ws['Q{}'.format(number + count)]

        promprim = ws['H{}'.format(number + count)]
        promseg = ws['M{}'.format(number + count)]
        promter = ws['R{}'.format(number + count)]

        orden.border = thin_border
        orden.alignment = Alignment(horizontal='center', vertical='center')
        name.border = thin_border
        name.alignment = Alignment(horizontal='left', vertical='center')
        dni.border = thin_border
        dni.alignment = Alignment(horizontal='center', vertical='center')
        nota1prim.border = thin_border
        nota1prim.alignment = Alignment(horizontal='center', vertical='center')
        nota1seg.border = thin_border
        nota1seg.alignment = Alignment(horizontal='center', vertical='center')
        nota1ter.border = thin_border
        nota1ter.alignment = Alignment(horizontal='center', vertical='center')

        nota2prim.border = thin_border
        nota2prim.alignment = Alignment(horizontal='center', vertical='center')
        nota2seg.border = thin_border
        nota2seg.alignment = Alignment(horizontal='center', vertical='center')
        nota2ter.border = thin_border
        nota2ter.alignment = Alignment(horizontal='center', vertical='center')

        nota3prim.border = thin_border
        nota3prim.alignment = Alignment(horizontal='center', vertical='center')
        nota3seg.border = thin_border
        nota3seg.alignment = Alignment(horizontal='center', vertical='center')
        nota3ter.border = thin_border
        nota3ter.alignment = Alignment(horizontal='center', vertical='center')

        nota4prim.border = thin_border
        nota4prim.alignment = Alignment(horizontal='center', vertical='center')
        nota4seg.border = thin_border
        nota4seg.alignment = Alignment(horizontal='center', vertical='center')
        nota4ter.border = thin_border
        nota4ter.alignment = Alignment(horizontal='center', vertical='center')

        promprim.border = thin_border
        promprim.alignment = Alignment(horizontal='center', vertical='center')
        promseg.border = thin_border
        promseg.alignment = Alignment(horizontal='center', vertical='center')
        promter.border = thin_border
        promter.alignment = Alignment(horizontal='center', vertical='center')

        ws['A{}'.format(number + count)] = "{}".format(count).zfill(2)
        ws['B{}'.format(number + count)] = "{}, {}".format(alumno.student.name, alumno.student.lastName).upper()
        ws['C{}'.format(number + count)] = "{:,}".format(alumno.student.dni).replace(',', '')
        ws['D{}'.format(number + count)] = "{}".format("")
        ws['I{}'.format(number + count)] = "{}".format("")
        ws['N{}'.format(number + count)] = "{}".format("")

        ws['E{}'.format(number + count)] = "{}".format("")
        ws['J{}'.format(number + count)] = "{}".format("")
        ws['O{}'.format(number + count)] = "{}".format("")

        ws['F{}'.format(number + count)] = "{}".format("")
        ws['K{}'.format(number + count)] = "{}".format("")
        ws['P{}'.format(number + count)] = "{}".format("")

        ws['H{}'.format(number + count)] = "{}".format("")
        ws['M{}'.format(number + count)] = "{}".format("")
        ws['R{}'.format(number + count)] = "{}".format("")

    workbook.save(response)
    return response
